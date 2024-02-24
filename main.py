if __name__ == '__main__':
    from openpyxl import Workbook
    from cri_modules import cri_main_index
    from cco_modules import cco_main_index
    import config
    import json
    import requests

    auth = config.auth

    # CRI WORKBOOK/WORKSHEET
    cri_workbook = Workbook()
    cri_work_sheet1 = cri_workbook.active
    cri_work_sheet2 = cri_workbook.create_sheet('Sheet1')

    # CCO WORKBOOK/WORKSHEET
    cco_workbook = Workbook()
    cco_work_sheet1 = cco_workbook.active
    cco_work_sheet2 = cco_workbook.create_sheet('Sheet1')

    # CRI DATA RESPONSE
    cri_url = config.cri_url
    response = requests.request("GET", cri_url, auth=auth)
    cri_data_issues = json.loads(response.text)

    url_issue_link = config.url_issue_link

    cri_worksheet = config.cri_ws

    labels = []
    item_list1 = []
    item_list2 = []

    status = 'Done'

    id_style = config.styleId
    style_id = id_style

    # CCO DATA RESPONSE
    cco_url = config.cco_url
    response = requests.request("GET", cco_url, auth=auth)
    cco_data_issues = json.loads(response.text)

    cco_worksheet = config.cco_ws

    def execCRIMainIndexData():
        cri_main_data = cri_main_index.__fetchCRIIssues__(cri_data_issues,
                                                          cri_worksheet,
                                                          item_list1,
                                                          item_list2,
                                                          labels,
                                                          style_id,
                                                          status,
                                                          url_issue_link,
                                                          cri_workbook,
                                                          cri_work_sheet1,
                                                          cri_work_sheet2)
        cri_main_data.displayData()
        cri_main_data.displayStyleSheet()


    def execCCOMainIndexData():
        cco_main_data = cco_main_index.__fetchCCOIssues__(cco_data_issues,
                                                          cco_workbook,
                                                          cco_worksheet,
                                                          cco_work_sheet1,
                                                          cco_work_sheet2)
        cco_main_data.displayData()
        cco_main_data.displayStyleSheet()

execCRIMainIndexData()
execCCOMainIndexData()
